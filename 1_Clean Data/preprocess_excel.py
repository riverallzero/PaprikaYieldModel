import datetime
import os.path
import openpyxl
import pandas as pd


def read_data(filename: str) -> None:
    doc = openpyxl.load_workbook(filename, data_only=True)
    df = pd.read_excel(filename)
    sheet = doc['환경 및 생산량 데이터']

    datelist = df.columns[8:]
    years = []
    months = []
    days = []
    datelist = datelist[:-8]
    for c in datelist:
        years.append(int(c.split('-')[0]))
        months.append(int(c.split('-')[1]))
        days.append(int(c.split('-')[2][:2]))

    farmlist = [0, 24, 48, 73, 97, 121, 145, 169, 193, 217, 241, 265, 286, 310, 334, 358, 382, 404, 428, 452, 475, 498, 522, 546]

    datasets = []

    for f, farm in enumerate(farmlist):
        for y, year in enumerate(years):
            row = farm + 2
            col = 3 + y + 6
            datasets.append({
                    "농가": sheet.cell(row=row, column=1).value,
                    "품종": sheet.cell(row=row, column=2).value,
                    "색상": sheet.cell(row=row, column=3).value,
                    "면적": sheet.cell(row=row, column=5).value,
                    "온실종류": sheet.cell(row=row, column=6).value,
                    "정식일자": sheet.cell(row=row, column=7).value,
                    "date": datelist[y][:-4],
                    "year": year,
                    "month": months[y],
                    "day": days[y],
                    "정식기준주차": (datetime.date(year, months[y], days[y])
                            - (datetime.date(2019, int(sheet.cell(row=row, column=7).value.split('-')[1]),
                                             int(sheet.cell(row=row, column=7).value.split('-')[2])))).days // 7,
                    "외부누적광량": sheet.cell(row=row, column=col).value,
                    "양액공급시작광": sheet.cell(row=row + 1, column=col).value,
                    "스라브EC주사기": sheet.cell(row=row + 2, column=col).value,
                    "스라브PH": sheet.cell(row=row + 3, column=col).value,
                    "공급횟수": sheet.cell(row=row + 4, column=col).value,
                    "1회공급량": sheet.cell(row=row + 5, column=col).value,
                    "드리퍼당공급량": sheet.cell(row=row + 6, column=col).value,
                    "공급총량": sheet.cell(row=row + 7, column=col).value,
                    "배액량": sheet.cell(row=row + 8, column=col).value,
                    "작물흡수량": sheet.cell(row=row + 10, column=col).value,
                    "광량당수분공급량": sheet.cell(row=row + 11, column=col).value,
                    "내부주간평균온도": sheet.cell(row=row + 12, column=col).value,
                    "내부야간평균온도": sheet.cell(row=row + 13, column=col).value,
                    "내부24시간평균온도": sheet.cell(row=row + 14, column=col).value,
                    "내부주간평균HD": sheet.cell(row=row + 15, column=col).value,
                    "내부야간평균HD": sheet.cell(row=row + 16, column=col).value,
                    "내부24시간평균HD": sheet.cell(row=row + 17, column=col).value,
                    "주간CO2": sheet.cell(row=row + 18, column=col).value,
                    "내부CO2": sheet.cell(row=row + 19, column=col).value,
                    "온실수확박스수": sheet.cell(row=row + 20, column=col).value,
                    "온실총생산량": sheet.cell(row=row + 21, column=col).value,
                    "단위면적당생산량": sheet.cell(row=row + 22, column=col).value,
                    "단위면적당누적생산량": sheet.cell(row=row + 23, column=col).value
            })

    return pd.DataFrame(datasets)


def main():
    output_dir = "../Output/Data"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    filename = "../Input/파프리카 농가 생산량 알고리즘 개발 자료(동계작형_정리본).xlsx"

    data = read_data(filename)
    output_filename = os.path.join(output_dir, "paprika_data.csv")
    data.to_csv(output_filename, index=False, encoding="utf-8-sig")


if __name__ == "__main__":
    main()